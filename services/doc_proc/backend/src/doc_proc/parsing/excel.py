"""Streaming Excel/CSV parser using openpyxl read_only mode.

Optimized for large spreadsheets using read-only streaming.
КСР.xlsx: 174K rows parsed in ~6s vs 80+ min with Docling.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from typing import Any

from doc_proc.models import ParsedElement, ParseResult

logger = logging.getLogger(__name__)

SECTION_KEYWORDS = re.compile(
    r"^\s*(Шифр|Раздел|Итого|Всего|Система|Подраздел|Глава|ИТОГО|ВСЕГО)",
    re.IGNORECASE,
)
SUBTOTAL_KEYWORDS = re.compile(
    r"(Итого|Всего|ИТОГО|ВСЕГО|итого\s+по)", re.IGNORECASE
)
SKIP_SHEET_PREFIXES = ("Source", "SmtRes", "Etalon", "Src")

HEADER_KEYWORDS: dict[str, list[str]] = {
    "number": ["№", "п/п", "номер", "n п/п", "№ п/п"],
    "code": ["обоснование", "обосно-", "шифр", "код", "код ресурса"],
    "name": [
        "наименование", "наим.", "наименование материала",
        "наименование работ", "наименование работ и затрат",
    ],
    "unit": ["единица", "ед. изм", "ед.", "измере-", "измерения"],
    "quantity": ["количество", "кол-во", "кол.", "коли-"],
    "note": ["примечание", "прим.", "примеч."],
}


class ExcelParser:
    """Streaming parser for Excel (xlsx/xls) and CSV files."""

    def can_handle(self, filename: str, content: bytes | None = None) -> bool:
        ext = filename.rsplit(".", 1)[-1].lower()
        return ext in ("xlsx", "xls", "csv")

    def parse(self, content: bytes, filename: str) -> ParseResult:
        ext = filename.rsplit(".", 1)[-1].lower()
        if ext == "csv":
            return self._parse_csv(content, filename)
        if ext == "xls":
            return self._parse_xls_legacy(content, filename)
        return self._parse_excel(content, filename)

    def parse_from_path(self, path: str, filename: str) -> ParseResult:
        with open(path, "rb") as f:
            return self.parse(f.read(), filename)

    def _parse_excel(self, content: bytes, filename: str) -> ParseResult:
        import openpyxl

        wb = openpyxl.load_workbook(
            io.BytesIO(content), read_only=True, data_only=True
        )
        all_elements: list[ParsedElement] = []
        element_idx = 0

        try:
            for ws in wb.worksheets:
                if any(ws.title.startswith(p) for p in SKIP_SHEET_PREFIXES):
                    continue

                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue

                header_idx, col_mapping = self._detect_header_row(rows)

                data_start = (header_idx + 1) if header_idx is not None else 0
                current_section = ws.title

                for i, row in enumerate(rows[data_start:], start=data_start):
                    cells = [str(c).strip() if c is not None else "" for c in row]
                    if all(not c for c in cells):
                        continue

                    row_type = self._classify_row(cells)
                    if row_type == "column_number":
                        continue

                    attrs = self._build_attributes(cells, col_mapping)

                    if row_type == "section_header":
                        current_section = next(
                            (c for c in cells if c and len(c) > 2), current_section
                        )

                    raw_text = self._build_semantic_raw_text(attrs, cells)
                    if not raw_text.strip():
                        continue

                    all_elements.append(ParsedElement(
                        index=element_idx,
                        text=raw_text,
                        element_type="data" if row_type == "data" else "section_header",
                        attributes=attrs,
                        section=current_section,
                        row_type=row_type,
                    ))
                    element_idx += 1
        finally:
            wb.close()

        return ParseResult(
            elements=all_elements,
            format_hint="tabular",
            document_type="spreadsheet",
            metadata={"parser": "excel_streaming", "filename": filename},
        )

    def _detect_header_row(
        self, rows: list[tuple], max_scan: int = 50
    ) -> tuple[int | None, dict[str, int]]:
        """Scan first N rows to find the header row by matching keywords."""
        best_score = 0
        best_idx = None
        best_mapping: dict[str, int] = {}

        for i, row in enumerate(rows[:max_scan]):
            score, mapping = self._match_header_row(row)
            if score > best_score:
                best_score = score
                best_idx = i
                best_mapping = mapping

        if best_score < 2:
            return None, {}
        return best_idx, best_mapping

    def _match_header_row(self, row: tuple) -> tuple[int, dict[str, int]]:
        """Score a row against header keywords. Returns (score, col_mapping)."""
        score = 0
        mapping: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_text = str(cell).strip().lower()
            if not cell_text:
                continue
            for field_name, keywords in HEADER_KEYWORDS.items():
                for kw in keywords:
                    if kw.lower() in cell_text or cell_text.startswith(kw.lower()):
                        if field_name not in mapping:
                            mapping[field_name] = col_idx
                            score += 1
                        break
        return score, mapping

    def _classify_row(self, cells: list[str]) -> str:
        """Classify row as data, section_header, subtotal, or column_number."""
        non_empty = [c for c in cells if c]
        if len(non_empty) < 2:
            text = " ".join(non_empty)
            if SECTION_KEYWORDS.search(text):
                return "section_header"
            if SUBTOTAL_KEYWORDS.search(text):
                return "subtotal"
            return "data"

        # All short integers → column numbering row
        if all(re.match(r"^\d{1,2}$", c) for c in non_empty):
            return "column_number"

        joined = " ".join(non_empty)
        if SUBTOTAL_KEYWORDS.search(joined):
            return "subtotal"
        if non_empty and SECTION_KEYWORDS.match(non_empty[0]):
            return "section_header"

        return "data"

    def _build_attributes(
        self, cells: list[str], col_mapping: dict[str, int]
    ) -> dict[str, Any]:
        """Map cells to named attributes using header mapping."""
        attrs: dict[str, Any] = {}
        for field_name, col_idx in col_mapping.items():
            if col_idx < len(cells) and cells[col_idx]:
                attrs[field_name] = cells[col_idx]
        # Add unmapped cells
        mapped_cols = set(col_mapping.values())
        for i, cell in enumerate(cells):
            if i not in mapped_cols and cell:
                attrs[f"col_{i}"] = cell
        return attrs

    def _build_semantic_raw_text(
        self, attrs: dict[str, Any], cells: list[str]
    ) -> str:
        """Build clean text optimized for embedding quality.

        Priority: name first, then compact code/unit/quantity.
        """
        name = attrs.get("name", "")
        code = attrs.get("code", "")
        unit = attrs.get("unit", "")
        quantity = attrs.get("quantity", "")

        if name:
            parts = [str(name)]
            if code:
                parts.append(f"({code})")
            if unit and quantity:
                parts.append(f"- {quantity} {unit}")
            elif unit:
                parts.append(f"[{unit}]")
            return " ".join(parts)

        # Fallback: join non-internal values
        internal_prefixes = ("col_", "number")
        values = [
            str(v) for k, v in attrs.items()
            if not any(k.startswith(p) for p in internal_prefixes) and v
        ]
        if values:
            return " ".join(values)

        # Ultimate fallback
        return " ".join(c for c in cells if c and len(c) > 1)

    def _parse_xls_legacy(self, content: bytes, filename: str) -> ParseResult:
        """Parse legacy .xls files by converting to .xlsx first via xlrd + openpyxl."""
        try:
            import xlrd

            xls_book = xlrd.open_workbook(file_contents=content)
            from openpyxl import Workbook

            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            for sheet_idx in range(xls_book.nsheets):
                xls_sheet = xls_book.sheet_by_index(sheet_idx)
                ws = wb.create_sheet(title=xls_sheet.name)
                for row_idx in range(xls_sheet.nrows):
                    for col_idx in range(xls_sheet.ncols):
                        ws.cell(
                            row=row_idx + 1,
                            column=col_idx + 1,
                            value=xls_sheet.cell_value(row_idx, col_idx),
                        )

            # Save to bytes and re-parse as xlsx
            buf = io.BytesIO()
            wb.save(buf)
            wb.close()
            buf.seek(0)
            return self._parse_excel(buf.read(), filename)
        except ImportError:
            logger.warning("xlrd not installed, falling back to Docling for .xls: %s", filename)
            raise ValueError(
                f"Cannot parse .xls file '{filename}': install xlrd package "
                f"(pip install xlrd) or convert to .xlsx"
            )

    def _parse_csv(self, content: bytes, filename: str) -> ParseResult:
        """Parse CSV with encoding detection."""
        import chardet

        detected = chardet.detect(content)
        encoding = detected.get("encoding", "utf-8") or "utf-8"

        for enc in (encoding, "utf-8", "cp1251", "latin-1"):
            try:
                text = content.decode(enc)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        else:
            text = content.decode("latin-1")

        reader = csv.reader(io.StringIO(text))
        rows_list = list(reader)
        if not rows_list:
            return ParseResult(
                elements=[], format_hint="tabular", document_type="csv"
            )

        headers = rows_list[0]
        elements: list[ParsedElement] = []
        current_section = ""

        for i, row in enumerate(rows_list[1:], start=1):
            attrs = {
                headers[j]: cell
                for j, cell in enumerate(row)
                if j < len(headers) and cell.strip()
            }
            text = " ".join(cell for cell in row if cell.strip())
            if not text.strip():
                continue

            non_empty = [c for c in row if c.strip()]
            if len(non_empty) <= 2 and SECTION_KEYWORDS.search(text):
                current_section = text.strip()
                el_type = "section_header"
            else:
                el_type = "data"

            elements.append(ParsedElement(
                index=i - 1,
                text=text,
                element_type=el_type,
                attributes=attrs,
                section=current_section,
                row_type="section_header" if el_type == "section_header" else "data",
            ))

        return ParseResult(
            elements=elements,
            format_hint="tabular",
            document_type="csv",
            metadata={"parser": "csv", "encoding": encoding},
        )
