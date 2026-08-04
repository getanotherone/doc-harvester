"""Bounded streaming XLSX extractor."""

from __future__ import annotations

from datetime import date, datetime, time
from io import BytesIO
from urllib.parse import urlsplit
from zipfile import ZipFile

from chunker import normalize_text
from doc_harvester.core import ContentBlock, ExtractedDocument, Extractor, FetchedArtifact


class _XLSXValidationError(Exception):
    pass


class XLSXExtractor(Extractor):
    """Stream workbook rows into sheet-scoped table blocks without evaluating formulas."""

    name = "xlsx"
    _MEDIA_TYPES = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    _CONTENT_TYPES_PATH = "[Content_Types].xml"
    _WORKBOOK_PATH = "xl/workbook.xml"

    def __init__(
        self,
        *,
        max_sheets: int = 100,
        max_rows: int = 200_000,
        max_cells: int = 2_000_000,
        max_entries: int = 5_000,
        max_uncompressed_bytes: int = 250 * 1024 * 1024,
        include_hidden_sheets: bool = False,
    ) -> None:
        for name, value in (
            ("XLSX max sheets", max_sheets),
            ("XLSX max rows", max_rows),
            ("XLSX max cells", max_cells),
            ("XLSX max entries", max_entries),
            ("XLSX max uncompressed bytes", max_uncompressed_bytes),
        ):
            if value < 1:
                raise ValueError(f"{name} must be at least 1")
        self.max_sheets = max_sheets
        self.max_rows = max_rows
        self.max_cells = max_cells
        self.max_entries = max_entries
        self.max_uncompressed_bytes = max_uncompressed_bytes
        self.include_hidden_sheets = include_hidden_sheets

    def supports(self, artifact: FetchedArtifact) -> bool:
        media_type = artifact.media_type.split(";", 1)[0].strip().lower()
        candidate = artifact.filename or urlsplit(artifact.resource.uri).path
        return media_type in self._MEDIA_TYPES or candidate.lower().endswith(".xlsx")

    def extract(self, artifact: FetchedArtifact) -> ExtractedDocument:
        if not self.supports(artifact):
            raise ValueError(f"{self.name} extractor does not support this artifact")
        if not artifact.content.startswith(b"PK"):
            raise ValueError("invalid XLSX container signature")

        try:
            self._validate_container(artifact.content)
            blocks, counts = self._extract_workbook(artifact.content)
        except _XLSXValidationError as error:
            raise ValueError(str(error)) from None
        except Exception as error:
            raise ValueError(f"XLSX extraction failed: {type(error).__name__}") from None

        return ExtractedDocument(
            artifact.resource,
            tuple(blocks),
            metadata={
                "extractor": self.name,
                "filename": artifact.filename,
                "media_type": artifact.media_type,
                **counts,
            },
        )

    def _validate_container(self, content: bytes) -> None:
        with ZipFile(BytesIO(content)) as archive:
            entries = archive.infolist()
            if len(entries) > self.max_entries:
                raise _XLSXValidationError(
                    f"XLSX exceeds configured entry limit ({self.max_entries})"
                )
            if sum(entry.file_size for entry in entries) > self.max_uncompressed_bytes:
                raise _XLSXValidationError(
                    "XLSX exceeds configured uncompressed-byte limit "
                    f"({self.max_uncompressed_bytes})"
                )
            if any(entry.flag_bits & 0x1 for entry in entries):
                raise _XLSXValidationError("encrypted XLSX entries are not supported")

            names = [entry.filename for entry in entries]
            if self._CONTENT_TYPES_PATH not in names or self._WORKBOOK_PATH not in names:
                raise _XLSXValidationError("XLSX is missing required OOXML parts")
            if names.count(self._WORKBOOK_PATH) != 1:
                raise _XLSXValidationError("XLSX contains duplicate workbook parts")

    def _extract_workbook(
        self, content: bytes
    ) -> tuple[list[ContentBlock], dict[str, object]]:
        from openpyxl import load_workbook

        workbook = load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        blocks: list[ContentBlock] = []
        total_rows = total_cells = nonempty_rows = formula_count = 0
        processed_sheets = 0
        skipped_hidden_sheet_count = 0
        try:
            worksheets = workbook.worksheets
            sheet_count = len(worksheets)
            if len(worksheets) > self.max_sheets:
                raise _XLSXValidationError(
                    f"XLSX exceeds configured sheet limit ({self.max_sheets})"
                )

            for sheet_index, worksheet in enumerate(worksheets):
                if worksheet.sheet_state != "visible" and not self.include_hidden_sheets:
                    skipped_hidden_sheet_count += 1
                    continue
                processed_sheets += 1

                declared_rows = worksheet.max_row or 0
                declared_columns = worksheet.max_column or 0
                if total_rows + declared_rows > self.max_rows:
                    raise _XLSXValidationError(
                        f"XLSX exceeds configured row limit ({self.max_rows})"
                    )
                if total_cells + (declared_rows * declared_columns) > self.max_cells:
                    raise _XLSXValidationError(
                        f"XLSX exceeds configured cell limit ({self.max_cells})"
                    )

                for row_index, row in enumerate(worksheet.iter_rows(), start=1):
                    total_rows += 1
                    if total_rows > self.max_rows:
                        raise _XLSXValidationError(
                            f"XLSX exceeds configured row limit ({self.max_rows})"
                        )
                    total_cells += len(row)
                    if total_cells > self.max_cells:
                        raise _XLSXValidationError(
                            f"XLSX exceeds configured cell limit ({self.max_cells})"
                        )

                    values = [cell.value for cell in row]
                    while values and values[-1] is None:
                        values.pop()
                    if not any(value is not None and str(value).strip() for value in values):
                        continue
                    nonempty_rows += 1

                    row_formula_count = sum(
                        cell.data_type == "f" for cell in row[: len(values)]
                    )
                    formula_count += row_formula_count
                    rendered = [self._render_value(value) for value in values]
                    blocks.append(
                        ContentBlock(
                            " | ".join(item.replace("|", "\\|") for item in rendered),
                            kind="table",
                            section=worksheet.title,
                            metadata={
                                "sheet": worksheet.title,
                                "sheet_index": sheet_index,
                                "sheet_state": worksheet.sheet_state,
                                "row": row_index,
                                "columns": len(values),
                                "non_empty_cells": sum(bool(item) for item in rendered),
                                "formula_cells": row_formula_count,
                            },
                        )
                    )
        finally:
            workbook.close()

        return blocks, {
            "sheet_count": sheet_count,
            "processed_sheet_count": processed_sheets,
            "skipped_hidden_sheet_count": skipped_hidden_sheet_count,
            "row_count": total_rows,
            "nonempty_row_count": nonempty_rows,
            "cell_count": total_cells,
            "block_count": len(blocks),
            "formula_count": formula_count,
            "formulas_evaluated": False,
        }

    @staticmethod
    def _render_value(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, (datetime, date, time)):
            return value.isoformat()
        return normalize_text(str(value))
